import json
import os

import openpyxl

SCORING_KEY_PATH = os.path.join(os.path.dirname(__file__), "..", "scoring_key.json")
T_SCORE_TABLE_PATH = os.path.join(os.path.dirname(__file__), "..", "T-score MMPI-2.xlsx")
SUBSCALE_T_SCORE_PATH = os.path.join(os.path.dirname(__file__), "..", "Subscale T1 score.xlsx")
SI_SUBSCALE_T_SCORE_PATH = os.path.join(os.path.dirname(__file__), "..", "Si Subscale T-conversion.xlsx")
SUPPLEMENTARY_T_SCORE_PATH = os.path.join(os.path.dirname(__file__), "..", "Supplementary scales T score.xlsx")

# Excel column headers per scale and gender for main scales
T_SCORE_COLUMNS = {
    "Male": {
        "L": "L (Male)",
        "F": "F(Male)",
        "K": "K(Male)",
        "1_Hs": "1-Hs(Male)",
        "2_D": "2-D(Male)",
        "3_Hy": "3-Hy(Male)",
        "4_Pd": "4-Pd(Male)",
        "5_Mf": "5-Mf(Male)",
        "6_Pa": "6-Pa(Male)",
        "7_Pt": "7-Pt(Male)",
        "8_Sc": "8-Sc(Male)",
        "9_Ma": "9-Ma(Male)",
        "0_Si": "0-Si(Male)",
    },
    "Female": {
        "L": "L (Female)",
        "F": "F(Female)",
        "K": "K(Female)",
        "1_Hs": "1-Hs(Female)",
        "2_D": "2-D(Female)",
        "3_Hy": "3-Hy(Female)",
        "4_Pd": "4-Pd(Female)",
        "5_Mf": "5-Mf(Female)",
        "6_Pa": "6-Pa(Female)",
        "7_Pt": "7-Pt(Female)",
        "8_Sc": "8-Sc(Female)",
        "9_Ma": "9-Ma(Female)",
        "0_Si": "0-Si(Female)",
    },
}

# Harris-Lingoes subscale column mappings
HARRIS_LINGOES_COLUMNS = {
    "Male": {
        "D1": "D1 (Men)", "D2": "D2 (Men)", "D3": "D3 (Men)", "D4": "D4 (Men)", "D5": "D5 (Men)",
        "Hy1": "Hy1 (Men)", "Hy2": "Hy2 (Men)", "Hy3": "Hy3 (Men)", "Hy4": "Hy4 (Men)", "Hy5": "Hy5 (Men)",
        "Pd1": "Pd1 (Men)", "Pd2": "Pd2 (Men)", "Pd3": "Pd3 (Men)", "Pd4": "Pd4 (Men)", "Pd5": "Pd5 (Men)",
        "Pa1": "Pa1 (Men)", "Pa2": "Pa2 (Men)", "Pa3": "Pa3 (Men)",
        "Sc1": "Sc1 (Men)", "Sc2": "Sc2 (Men)", "Sc3": "Sc3 (Men)", "Sc4": "Sc4 (Men)", "Sc5": "Sc5 (Men)", "Sc6": "Sc6 (Men)",
        "Ma1": "Ma1 (Men)", "Ma2": "Ma2 (Men)", "Ma3": "Ma3 (Men)", "Ma4": "Ma4 (Men)",
    },
    "Female": {
        "D1": "D1 (Women)", "D2": "D2 (Women)", "D3": "D3 (Women)", "D4": "D4 (Women)", "D5": "D5 (Women)",
        "Hy1": "Hy1 (Women)", "Hy2": "Hy2 (Women)", "Hy3": "Hy3 (Women)", "Hy4": "Hy4 (Women)", "Hy5": "Hy5 (Women)",
        "Pd1": "Pd1 (Women)", "Pd2": "Pd2 (Women)", "Pd3": "Pd3 (Women)", "Pd4": "Pd4 (Women)", "Pd5": "Pd5 (Women)",
        "Pa1": "Pa1 (Women)", "Pa2": "Pa2 (Women)", "Pa3": "Pa3 (Women)",
        "Sc1": "Sc1 (Women)", "Sc2": "Sc2 (Women)", "Sc3": "Sc3 (Women)", "Sc4": "Sc4 (Women)", "Sc5": "Sc5 (Women)", "Sc6": "Sc6 (Women)",
        "Ma1": "Ma1 (Women)", "Ma2": "Ma2 (Women)", "Ma3": "Ma3 (Women)", "Ma4": "Ma4 (Women)",
    },
}

# Si subscale column mappings
SI_SUBSCALE_COLUMNS = {
    "Male": {"Si1": "Si1 (Men)", "Si2": "Si2 (Men)", "Si3": "Si3 (Men)"},
    "Female": {"Si1": "Si1 (Women)", "Si2": "Si2 (Women)", "Si3": "Si3 (Women)"},
}

# Supplementary scale column mappings (note: Fb column has typo "Fb (Men" without closing paren)
# OH, Do, Re, Mt, GM, GF, PK, PS use Male/Female instead of Men/Women in Excel headers
SUPPLEMENTARY_T_COLUMNS = {
    "Male": {
        "A": "A (Men)", "R": "R (Men)", "Es": "Es (Men)", "MAC-R": "MAC-R (Men)",
        "Fb": "Fb (Men", "TRIN": "TRIN (Men)", "VRIN": "VRIN (Men)",
        "OH": "O-H (Male)", "Do": "Do (Male)", "Re": "Re (Male)", "Mt": "Mt (Male)",
        "GM": "GM (Male)", "GF": "GF (Male)", "PK": "PK (Male)", "PS": "PS (Male)",
    },
    "Female": {
        "A": "A (Women)", "R": "R (Women)", "Es": "Es (Women)", "MAC-R": "MAC-R (Women)",
        "Fb": "Fb (Women)", "TRIN": "TRIN (Women)", "VRIN": "VRIN (Women)",
        "OH": "O-H (Female)", "Do": "Do (Female)", "Re": "Re (Female)", "Mt": "Mt (Female)",
        "GM": "GM (Female)", "GF": "GF (Female)", "PK": "PK (Female)", "PS": "PS (Female)",
    },
}

_t_score_tables: dict[str, dict[int, int]] | None = None
_subscale_t_tables: dict[str, dict[int, int]] | None = None
_si_subscale_t_tables: dict[str, dict[int, int]] | None = None
_supplementary_t_tables: dict[str, dict[int, int]] | None = None

# Map JSON scale keys to the canonical norm keys used throughout the app
SCALE_NORM_MAP = {
    "1_Hs": "1_Hs",
    "2_D": "2_D",
    "3_Hy": "3_Hy",
    "4_Pd": "4_Pd",
    "5_Mf_male": "5_Mf",
    "5_Mf_female": "5_Mf",
    "6_Pa": "6_Pa",
    "7_Pt": "7_Pt",
    "8_Sc": "8_Sc",
    "9_Ma": "9_Ma",
    "0_Si": "0_Si",
}


def load_scoring_key() -> dict:
    with open(SCORING_KEY_PATH, "r") as f:
        data = json.load(f)
    return data.get("MMPI_scoring_key", data)


def load_t_score_tables() -> dict[str, dict[int, int]]:
    """Load raw-score → T-score lookup tables from T-score MMPI-2.xlsx."""
    global _t_score_tables
    if _t_score_tables is not None:
        return _t_score_tables

    wb = openpyxl.load_workbook(T_SCORE_TABLE_PATH, data_only=True)
    ws = wb.active

    header_to_col: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header and header != "Raw score ":
            header_to_col[str(header).strip()] = col

    tables: dict[str, dict[int, int]] = {}
    for gender, scale_cols in T_SCORE_COLUMNS.items():
        for scale, col_name in scale_cols.items():
            col = header_to_col.get(col_name)
            if col is None:
                raise ValueError(f"T-score column not found in Excel: {col_name}")
            key = f"{gender}:{scale}"
            tables[key] = {}
            for row in range(2, ws.max_row + 1):
                raw = ws.cell(row, 1).value
                t_val = ws.cell(row, col).value
                if raw is not None and t_val is not None:
                    tables[key][int(raw)] = int(t_val)

    _t_score_tables = tables
    return tables


def load_subscale_t_tables() -> dict[str, dict[int, int]]:
    """Load Harris-Lingoes subscale T-score tables from Subscale T1 score.xlsx."""
    global _subscale_t_tables
    if _subscale_t_tables is not None:
        return _subscale_t_tables

    wb = openpyxl.load_workbook(SUBSCALE_T_SCORE_PATH, data_only=True)
    ws = wb.active

    header_to_col: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header:
            header_to_col[str(header).strip()] = col

    tables: dict[str, dict[int, int]] = {}
    for gender, scale_cols in HARRIS_LINGOES_COLUMNS.items():
        for scale, col_name in scale_cols.items():
            col = header_to_col.get(col_name)
            if col is None:
                continue
            key = f"{gender}:{scale}"
            tables[key] = {}
            for row in range(2, ws.max_row + 1):
                raw = ws.cell(row, 1).value
                t_val = ws.cell(row, col).value
                if raw is not None and t_val is not None:
                    tables[key][int(raw)] = int(t_val)

    _subscale_t_tables = tables
    return tables


def load_si_subscale_t_tables() -> dict[str, dict[int, int]]:
    """Load Si subscale T-score tables from Si Subscale T-conversion.xlsx."""
    global _si_subscale_t_tables
    if _si_subscale_t_tables is not None:
        return _si_subscale_t_tables

    wb = openpyxl.load_workbook(SI_SUBSCALE_T_SCORE_PATH, data_only=True)
    ws = wb.active

    header_to_col: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header:
            header_to_col[str(header).strip()] = col

    tables: dict[str, dict[int, int]] = {}
    for gender, scale_cols in SI_SUBSCALE_COLUMNS.items():
        for scale, col_name in scale_cols.items():
            col = header_to_col.get(col_name)
            if col is None:
                continue
            key = f"{gender}:{scale}"
            tables[key] = {}
            for row in range(2, ws.max_row + 1):
                raw = ws.cell(row, 1).value
                t_val = ws.cell(row, col).value
                if raw is not None and t_val is not None:
                    tables[key][int(raw)] = int(t_val)

    _si_subscale_t_tables = tables
    return tables


def load_supplementary_t_tables() -> dict[str, dict[int, int]]:
    """Load supplementary scale T-score tables from Supplementary scales T score.xlsx."""
    global _supplementary_t_tables
    if _supplementary_t_tables is not None:
        return _supplementary_t_tables

    wb = openpyxl.load_workbook(SUPPLEMENTARY_T_SCORE_PATH, data_only=True)
    ws = wb.active

    header_to_col: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(1, col).value
        if header:
            # Keep original header (may have typos like "Fb (Men")
            header_to_col[str(header).strip()] = col

    tables: dict[str, dict[int, int]] = {}
    for gender, scale_cols in SUPPLEMENTARY_T_COLUMNS.items():
        for scale, col_name in scale_cols.items():
            col = header_to_col.get(col_name)
            if col is None:
                # Try without trailing paren for typos
                col = header_to_col.get(col_name.rstrip(')'))
            if col is None:
                continue
            key = f"{gender}:{scale}"
            tables[key] = {}
            for row in range(2, ws.max_row + 1):
                raw = ws.cell(row, 1).value
                t_val = ws.cell(row, col).value
                if raw is not None and t_val is not None:
                    # Handle TRIN values like "114F", "57T" - extract just the number
                    if isinstance(t_val, str):
                        # Remove trailing T/F direction indicator
                        t_val = t_val.rstrip('TF')
                        try:
                            t_val = int(t_val)
                        except ValueError:
                            continue
                    tables[key][int(raw)] = int(t_val)

    _supplementary_t_tables = tables
    return tables


def _lookup_t_score(scale: str, raw_score: float, gender: str) -> int:
    """Look up T-score from Excel table; clamp raw score to table range."""
    tables = load_t_score_tables()
    gender_key = "Female" if gender == "Female" else "Male"
    table = tables[f"{gender_key}:{scale}"]
    raw_int = int(round(raw_score))
    min_raw, max_raw = min(table.keys()), max(table.keys())
    raw_int = max(min_raw, min(max_raw, raw_int))
    return int(table[raw_int])


def _lookup_subscale_t_score(scale: str, raw_score: int, gender: str, is_si: bool = False) -> int | None:
    """Look up subscale T-score from appropriate Excel table."""
    gender_key = "Female" if gender == "Female" else "Male"
    key = f"{gender_key}:{scale}"
    
    if is_si:
        tables = load_si_subscale_t_tables()
    else:
        tables = load_subscale_t_tables()
    
    if key not in tables:
        return None
    
    table = tables[key]
    if not table:
        return None
    
    raw_int = int(round(raw_score))
    min_raw, max_raw = min(table.keys()), max(table.keys())
    raw_int = max(min_raw, min(max_raw, raw_int))
    return int(table.get(raw_int, 50))


def _lookup_supplementary_t_score(scale: str, raw_score: int, gender: str) -> int | None:
    """Look up supplementary scale T-score from Excel table."""
    gender_key = "Female" if gender == "Female" else "Male"
    key = f"{gender_key}:{scale}"
    
    tables = load_supplementary_t_tables()
    
    if key not in tables:
        return None
    
    table = tables[key]
    if not table:
        return None
    
    raw_int = int(round(raw_score))
    min_raw, max_raw = min(table.keys()), max(table.keys())
    raw_int = max(min_raw, min(max_raw, raw_int))
    return int(table.get(raw_int, 50))


def _count_scale(answers: dict[int, bool], scale_def: dict) -> int:
    """Count raw score for a scale given answer map {question_number: True/False}."""
    score = 0
    for item in scale_def.get("true_items", []):
        if answers.get(item) is True:
            score += 1
    for item in scale_def.get("false_items", []):
        if answers.get(item) is False:
            score += 1
    return score


def _count_vrin(answers: dict[int, bool], scale_def: dict) -> int:
    """Count VRIN score - pairs where responses match the specified pattern indicate inconsistency."""
    score = 0
    for pair in scale_def.get("pairs", []):
        q1, resp1, q2, resp2 = pair
        expected1 = resp1 == "T"
        expected2 = resp2 == "T"
        actual1 = answers.get(q1)
        actual2 = answers.get(q2)
        if actual1 is not None and actual2 is not None:
            if actual1 == expected1 and actual2 == expected2:
                score += 1
    return score


def _count_trin(answers: dict[int, bool], scale_def: dict) -> tuple[int, str]:
    """
    Count TRIN score.
    Returns (raw_score, direction) where direction is 'T' or 'F'.
    True pairs: both answered True = acquiescent (True) responding
    False pairs: both answered False = non-acquiescent (False) responding
    Final score = true_count - false_count + base_score
    Direction is T if true_count > false_count, F otherwise
    """
    true_count = 0
    false_count = 0
    
    # True pairs: both answered True = acquiescent responding
    for pair in scale_def.get("true_pairs", []):
        q1, q2 = pair
        if answers.get(q1) is True and answers.get(q2) is True:
            true_count += 1
    
    # False pairs: both answered False = non-acquiescent responding
    for pair in scale_def.get("false_pairs", []):
        q1, q2 = pair
        if answers.get(q1) is False and answers.get(q2) is False:
            false_count += 1
    
    # Base score to avoid negatives
    base = scale_def.get("base_score", 9)
    raw_score = true_count - false_count + base
    
    # Direction indicates whether True or False responding predominates
    direction = "T" if true_count >= false_count else "F"
    
    return raw_score, direction


def calculate_raw_scores(answers: dict[int, bool], gender: str) -> dict[str, float]:
    key = load_scoring_key()
    raw = {}

    for scale_name, scale_def in key["validity_scales"].items():
        raw[scale_name] = _count_scale(answers, scale_def)

    mf_key = "5_Mf_female" if gender == "Female" else "5_Mf_male"

    for scale_name, scale_def in key["clinical_scales"].items():
        # Pick the gender-appropriate Mf scale, skip the other
        if scale_name in ("5_Mf_male", "5_Mf_female"):
            if scale_name != mf_key:
                continue
            raw["5_Mf"] = _count_scale(answers, scale_def)
        else:
            raw[scale_name] = _count_scale(answers, scale_def)

    return raw


def calculate_subscale_scores(answers: dict[int, bool], gender: str) -> dict[str, dict]:
    """Calculate raw scores and T-scores for Harris-Lingoes, Si, and Supplementary scales."""
    key = load_scoring_key()
    
    result = {
        "harris_lingoes": {},
        "si_subscales": {},
        "supplementary": {},
    }
    
    # Harris-Lingoes subscales with T-scores
    if "harris_lingoes_subscales" in key:
        for scale_name, scale_def in key["harris_lingoes_subscales"].items():
            raw = _count_scale(answers, scale_def)
            t_score = _lookup_subscale_t_score(scale_name, raw, gender, is_si=False)
            result["harris_lingoes"][scale_name] = {
                "raw": raw,
                "t_score": t_score
            }
    
    # Si subscales with T-scores
    if "si_subscales" in key:
        for scale_name, scale_def in key["si_subscales"].items():
            raw = _count_scale(answers, scale_def)
            t_score = _lookup_subscale_t_score(scale_name, raw, gender, is_si=True)
            result["si_subscales"][scale_name] = {
                "raw": raw,
                "t_score": t_score
            }
    
    # Supplementary scales with T-scores
    if "supplementary_scales" in key:
        for scale_name, scale_def in key["supplementary_scales"].items():
            if scale_name == "VRIN":
                raw = _count_vrin(answers, scale_def)
                t_score = _lookup_supplementary_t_score("VRIN", raw, gender)
                result["supplementary"][scale_name] = {"raw": raw, "t_score": t_score}
            elif scale_name == "TRIN":
                raw, direction = _count_trin(answers, scale_def)
                t_score = _lookup_supplementary_t_score("TRIN", raw, gender)
                result["supplementary"][scale_name] = {
                    "raw": raw,
                    "t_score": t_score,
                    "direction": direction
                }
            else:
                raw = _count_scale(answers, scale_def)
                # Look up T-score for scales that have T-score tables
                t_score = _lookup_supplementary_t_score(scale_name, raw, gender)
                result["supplementary"][scale_name] = {"raw": raw, "t_score": t_score}
    
    return result


def apply_k_correction(raw_scores: dict[str, float], gender: str) -> dict[str, float]:
    key = load_scoring_key()
    raw_k = raw_scores.get("K", 0)
    corrected = dict(raw_scores)

    mf_key = "5_Mf_female" if gender == "Female" else "5_Mf_male"

    for scale_name, scale_def in key["clinical_scales"].items():
        if scale_name in ("5_Mf_male", "5_Mf_female") and scale_name != mf_key:
            continue

        factor = scale_def.get("k_correction_factor")
        if factor is None or factor == 0:
            continue

        norm_name = SCALE_NORM_MAP.get(scale_name, scale_name)
        if norm_name in corrected:
            corrected[norm_name] = round(corrected[norm_name] + (factor * raw_k))

    return corrected


def calculate_t_scores(raw_scores: dict[str, float], gender: str) -> dict[str, int]:
    """Convert raw scores (NOT K-corrected) to integer T-scores via T-score MMPI-2.xlsx."""
    return {
        scale_name: _lookup_t_score(scale_name, raw_val, gender)
        for scale_name, raw_val in raw_scores.items()
    }


def full_scoring_pipeline(answers: dict[int, bool], gender: str) -> dict:
    raw_scores = calculate_raw_scores(answers, gender)
    k_corrected = apply_k_correction(raw_scores, gender)
    t_scores = calculate_t_scores(raw_scores, gender)
    subscales = calculate_subscale_scores(answers, gender)

    return {
        "raw_scores": raw_scores,
        "k_corrected_scores": k_corrected,
        "t_scores": t_scores,
        "harris_lingoes_subscales": subscales["harris_lingoes"],
        "si_subscales": subscales["si_subscales"],
        "supplementary_scales": subscales["supplementary"],
    }
