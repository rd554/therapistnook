"""
Seed the database with 567 MMPI-2 questions and the owner account.

Usage:
    python3 seed.py                                      # Auto-detect xlsx or use placeholders
    python3 seed.py --xlsx path/to/Questionnaire.xlsx    # Load from xlsx
    python3 seed.py --csv  path/to/questions.csv         # Load from CSV
"""

import asyncio
import argparse
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))

from dotenv import load_dotenv
load_dotenv()

from database import engine, async_session, Base
from models import Question, Practitioner
from auth import hash_password, generate_ref_code

TOTAL_QUESTIONS = 567
XLSX_PATH_DEFAULT = os.path.join(os.path.dirname(__file__), "..", "MMPI - 2 - Questionnaire.xlsx")

OWNER_NAME = os.getenv("OWNER_NAME", "Admin")
OWNER_EMAIL = os.getenv("OWNER_EMAIL", "admin@mmpi.local")
OWNER_PASSWORD = os.getenv("OWNER_PASSWORD", "admin123")


def load_questions_from_xlsx(xlsx_path: str) -> dict[int, str]:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    questions = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, text = row[0], row[1]
        if isinstance(num, (int, float)) and text and 1 <= int(num) <= TOTAL_QUESTIONS:
            questions[int(num)] = str(text).strip()
    wb.close()
    return questions


def load_questions_from_csv(csv_path: str) -> dict[int, str]:
    questions = {}
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []

        num_col = next((c for c in cols if c.lower().strip() in (
            "number", "item_number", "item", "no", "#", "q_no", "question_number", "sr. no."
        )), None)
        text_col = next((c for c in cols if c.lower().strip() in (
            "text", "question", "question_text", "item_text", "content", "statement"
        )), None)

        if not num_col:
            for c in cols:
                try:
                    int(list(csv.DictReader(open(csv_path)))[0][c])
                    num_col = c
                    break
                except (ValueError, IndexError):
                    continue
        if not text_col:
            for c in cols:
                if c != num_col:
                    text_col = c
                    break

        if not num_col or not text_col:
            print(f"Could not detect columns. Found: {cols}")
            sys.exit(1)

        print(f"Using columns: number='{num_col}', text='{text_col}'")
        f.seek(0)
        reader = csv.DictReader(f)
        for row in reader:
            try:
                num = int(row[num_col].strip())
                text = row[text_col].strip()
                if 1 <= num <= TOTAL_QUESTIONS and text:
                    questions[num] = text
            except (ValueError, KeyError):
                continue
    return questions


def generate_placeholder_questions() -> dict[int, str]:
    return {i: f"MMPI-2 Item {i} — [Replace with actual question text from CSV]" for i in range(1, TOTAL_QUESTIONS + 1)}


async def seed(questions: dict[int, str]):
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with async_session() as session:
        from sqlalchemy import select, func, delete

        # ── Seed questions ───────────────────────────────────────────────────
        count = (await session.execute(select(func.count(Question.id)))).scalar() or 0
        if count >= TOTAL_QUESTIONS:
            print(f"Database already has {count} questions. Clearing and re-seeding...")
            await session.execute(delete(Question))
            await session.commit()

        for num in sorted(questions.keys()):
            session.add(Question(number=num, text=questions[num]))
        await session.commit()
        final_count = (await session.execute(select(func.count(Question.id)))).scalar()
        print(f"Seeded {final_count} questions.")

        # ── Seed owner account ───────────────────────────────────────────────
        existing = (await session.execute(
            select(Practitioner).where(Practitioner.email == OWNER_EMAIL)
        )).scalar_one_or_none()

        if existing:
            print(f"Owner account already exists: {OWNER_EMAIL} (ref: {existing.ref_code})")
        else:
            ref = generate_ref_code()
            owner = Practitioner(
                name=OWNER_NAME,
                email=OWNER_EMAIL,
                password_hash=hash_password(OWNER_PASSWORD),
                role="owner",
                ref_code=ref,
            )
            session.add(owner)
            await session.commit()
            print(f"Created owner account: {OWNER_EMAIL} / {OWNER_PASSWORD}")
            print(f"Owner ref code (test link): {ref}")


def main():
    parser = argparse.ArgumentParser(description="Seed MMPI-2 questions into database")
    parser.add_argument("--xlsx", type=str, help="Path to .xlsx file with questions")
    parser.add_argument("--csv", type=str, help="Path to .csv file with questions")
    args = parser.parse_args()

    if args.xlsx:
        path = args.xlsx
        if not os.path.exists(path):
            print(f"File not found: {path}"); sys.exit(1)
        print(f"Loading questions from xlsx: {path}")
        questions = load_questions_from_xlsx(path)
        print(f"Loaded {len(questions)} questions")
    elif args.csv:
        if not os.path.exists(args.csv):
            print(f"CSV file not found: {args.csv}"); sys.exit(1)
        print(f"Loading questions from CSV: {args.csv}")
        questions = load_questions_from_csv(args.csv)
        print(f"Loaded {len(questions)} questions")
    elif os.path.exists(XLSX_PATH_DEFAULT):
        print(f"Auto-detected questionnaire: {XLSX_PATH_DEFAULT}")
        questions = load_questions_from_xlsx(XLSX_PATH_DEFAULT)
        print(f"Loaded {len(questions)} questions")
    else:
        print("No CSV/XLSX provided. Using placeholder question text.")
        questions = generate_placeholder_questions()

    missing = set(range(1, TOTAL_QUESTIONS + 1)) - set(questions.keys())
    if missing:
        print(f"Warning: {len(missing)} questions missing, filling with placeholders")
        for num in missing:
            questions[num] = f"MMPI-2 Item {num} — [Question text not available]"

    asyncio.run(seed(questions))


if __name__ == "__main__":
    main()
